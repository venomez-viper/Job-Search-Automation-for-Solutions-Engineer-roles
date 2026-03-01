"""
onedrive.py — Append new job rows to an Excel file stored in Microsoft OneDrive.

Uses Microsoft Graph API with client credentials (app-only auth).
The seen_jobs.db deduplicator guarantees only new jobs are ever passed here,
so rows will never be duplicated in the Excel file.

Setup (one-time, ~5 min):
  1. portal.azure.com → Azure Active Directory → App registrations → New registration
     - Name: "Job Hunt Automation" (any name)
     - Supported account types: "Accounts in this organizational directory only"
  2. API Permissions → Add a permission → Microsoft Graph → Application permissions
     → Files.ReadWrite.All → Grant admin consent
  3. Certificates & Secrets → New client secret → copy the Value
  4. Copy: Application (client) ID, Directory (tenant) ID, Client Secret Value
  5. Create a folder called "Jobs" in your OneDrive and create/upload an Excel file
     (e.g. job_tracker.xlsx) — OR let this script create it automatically.
  6. Add these GitHub Secrets:
       AZURE_CLIENT_ID     = Application (client) ID
       AZURE_CLIENT_SECRET = Client Secret Value
       AZURE_TENANT_ID     = Directory (tenant) ID
       ONEDRIVE_FILE_PATH  = /Jobs/job_tracker.xlsx   (path in your OneDrive)

If any secret is missing, this logger is silently skipped.
"""
import os
import io
import logging
from datetime import datetime

from fetchers.base import Job

logger = logging.getLogger(__name__)

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
TOKEN_URL  = "https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
HEADERS    = ["Date", "Company", "Title", "Location", "URL", "Score", "Label", "Source", "Posted", "Match Signals"]


def _get_token(tenant_id: str, client_id: str, client_secret: str) -> str | None:
    """Fetch a Graph API access token using client credentials."""
    import requests
    resp = requests.post(
        TOKEN_URL.format(tenant=tenant_id),
        data={
            "grant_type":    "client_credentials",
            "client_id":     client_id,
            "client_secret": client_secret,
            "scope":         "https://graph.microsoft.com/.default",
        },
        timeout=20,
    )
    if resp.status_code != 200:
        logger.error(f"OneDrive: token request failed ({resp.status_code}): {resp.text[:200]}")
        return None
    return resp.json().get("access_token")


def _graph(method: str, path: str, token: str, **kwargs):
    """Make a Graph API call."""
    import requests
    return requests.request(
        method,
        f"{GRAPH_BASE}{path}",
        headers={"Authorization": f"Bearer {token}", **kwargs.pop("extra_headers", {})},
        timeout=30,
        **kwargs,
    )


def log_to_onedrive(jobs: list[Job], dry_run: bool = False) -> bool:
    """
    Append job rows to the OneDrive Excel file.
    Creates the file with a header row if it doesn't exist yet.
    """
    if dry_run:
        logger.info("DRY RUN — skipping OneDrive log.")
        return True

    tenant_id     = os.environ.get("AZURE_TENANT_ID", "")
    client_id     = os.environ.get("AZURE_CLIENT_ID", "")
    client_secret = os.environ.get("AZURE_CLIENT_SECRET", "")
    file_path     = os.environ.get("ONEDRIVE_FILE_PATH", "/Jobs/job_tracker.xlsx").lstrip("/")

    if not all([tenant_id, client_id, client_secret]):
        logger.info("OneDrive: Azure credentials not set — skipping. "
                    "(Add AZURE_CLIENT_ID, AZURE_CLIENT_SECRET, AZURE_TENANT_ID to GitHub Secrets)")
        return True

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("OneDrive: openpyxl not installed. Run: pip install openpyxl")
        return False

    try:
        import requests as req
        token = _get_token(tenant_id, client_id, client_secret)
        if not token:
            return False

        # ── Try to download existing file ────────────────────────────────
        drive_path = f"/me/drive/root:/{file_path}"
        dl_resp = _graph("GET", f"{drive_path}:/content", token)

        if dl_resp.status_code == 200:
            wb = openpyxl.load_workbook(io.BytesIO(dl_resp.content))
            ws = wb.active
            logger.info(f"OneDrive: opened existing file '{file_path}' ({ws.max_row} rows).")
        elif dl_resp.status_code in (404, 400):
            # File doesn't exist yet — create fresh
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Jobs"
            ws.append(HEADERS)
            # Style header row
            fill = PatternFill("solid", fgColor="1a73e8")
            font = Font(bold=True, color="FFFFFF")
            for col in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center")
            logger.info(f"OneDrive: creating new file '{file_path}'.")
        else:
            logger.error(f"OneDrive: unexpected response downloading file: {dl_resp.status_code}")
            return False

        # ── Append new rows ───────────────────────────────────────────────
        today = datetime.now().strftime("%Y-%m-%d")
        for job in jobs:
            bd = getattr(job, "breakdown", None)
            signals = ""
            if bd:
                parts = []
                if bd.core_matches:
                    parts.append("Core: " + ", ".join(m.split("(")[0] for m in bd.core_matches[:3]))
                if bd.tool_matches:
                    parts.append("Tools: " + ", ".join(m.split("(")[0] for m in bd.tool_matches[:2]))
                if bd.transferable_matches:
                    parts.append("Transfer: " + ", ".join(m.split("->")[0] for m in bd.transferable_matches[:2]))
                signals = " | ".join(parts)

            ws.append([
                today,
                job.company,
                job.title,
                job.location,
                job.url,
                job.score,
                getattr(job, "label", ""),
                job.source,
                job.date_posted or "",
                signals,
            ])

        # Auto-fit columns
        col_widths = [12, 22, 38, 20, 55, 7, 18, 12, 12, 60]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # ── Upload back to OneDrive ───────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        up_resp = _graph(
            "PUT",
            f"{drive_path}:/content",
            token,
            extra_headers={"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
            data=buf.read(),
        )
        if up_resp.status_code in (200, 201):
            logger.info(f"OneDrive: uploaded {len(jobs)} new rows to '{file_path}'.")
            return True
        else:
            logger.error(f"OneDrive: upload failed ({up_resp.status_code}): {up_resp.text[:300]}")
            return False

    except Exception as e:
        logger.error(f"OneDrive log failed: {e}")
        return False
